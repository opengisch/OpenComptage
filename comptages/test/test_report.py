from pathlib import Path
import os
import pytz
from datetime import datetime
from django.test import TransactionTestCase
from django.core.management import call_command
from openpyxl import load_workbook

from comptages.test import utils as test_utils
from comptages.test import yearly_count_for
from comptages.datamodel import models
from comptages.report.yearly_report_bike import YearlyReportBike
from comptages.core import report, importer


class ImportTest(TransactionTestCase):
    @classmethod
    def setUpClass(cls):
        cls.test_outputs = Path("/OpenComptage/test_outputs")
        cls.test_data = Path("/OpenComptage/test_data")

        for file in cls.test_outputs.iterdir():
            os.remove(file)

    def setUp(self):
        for file in Path(self.test_outputs).iterdir():
            os.remove(file)
        # With TransactionTestCase the db is reset at every test, so we
        # re-import base data every time.
        call_command("importdata")
        for file in Path(self.test_outputs).iterdir():
            os.remove(file)

    def test_report(self):
        # Create count and import some data
        installation = models.Installation.objects.get(name="00056520")
        count = yearly_count_for(2021, installation)
        importer.import_file(test_utils.test_data_path("00056520.V01"), count)
        importer.import_file(test_utils.test_data_path("00056520.V02"), count)
        report.prepare_reports("/tmp/", count)

    def test_all_sections_default(self):
        # Test if default report features all sections for special case
        test_data_folder = "5350_1_4"
        section_id = "53526896"

        installation = models.Installation.objects.get(lane__id_section_id=section_id)
        n_sections = (
            models.Lane.objects.filter(id_installation=installation.id)
            .values("id_section")
            .count()
        )
        self.assertGreater(n_sections, 0)

        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SWISS10")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 3, 1)),
            end_service_date=tz.localize(datetime(2021, 3, 2)),
            start_process_date=tz.localize(datetime(2021, 3, 15)),
            end_process_date=tz.localize(datetime(2021, 3, 28)),
            start_put_date=tz.localize(datetime(2021, 2, 20)),
            end_put_date=tz.localize(datetime(2021, 4, 1)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        importer.import_file(test_utils.test_data_path("00056520.V01"), count)
        importer.import_file(test_utils.test_data_path("00056520.V02"), count)

        report.prepare_reports("/tmp/", count)

    def test_busiest_by_season(self):
        # Import test data pertaining to "mobilité douce"
        installation = models.Installation.objects.get(name="00107695")
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = Path("/test_data").joinpath(
            "64540060_Latenium_PS2021_ChMixte.txt"
        )
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        # Collect count details
        details = YearlyReportBike.count_details_by_season(count)

        def inspect_leaves(d, res) -> list[int]:
            for v in d.values():
                if isinstance(v, int):
                    res.append(v)
                elif isinstance(v, dict):
                    inspect_leaves(v, res)
            return res

        self.assertTrue(all(value > 0 for value in inspect_leaves(details, [])))

        # Prepare workbook
        path_to_inputs = Path("comptages/report").joinpath("template_yearly_bike.xlsx")
        path_to_outputs = self.test_outputs.joinpath("yearly_bike.xlsx")
        wb = load_workbook(path_to_inputs)

        # Write data & save
        ws = wb["Data_yearly_stats"]
        window = ws["B22:F25"]

        for row, season in zip(window, ("printemps", "été", "automne", "hiver")):
            for cell, category in zip(
                row, ("VELO", "MONO", "SHORT", "SPECIAL", "MULTI")
            ):
                cell.value = sum(details[season][category].values())

        wb.save(path_to_outputs)

    def test_busiest_by_day_month(self):
        # Import test data pertaining to "mobilité douce"
        installation = models.Installation.objects.get(name="00107695")
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = Path("/test_data").joinpath(
            "64540060_Latenium_PS2021_ChMixte.txt"
        )
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        # Collecting count details
        data = YearlyReportBike.count_details_by_day_month(count)

        # Prepare workbook
        path_to_inputs = Path("comptages/report").joinpath("template_yearly_bike.xlsx")
        path_to_outputs = Path("/test_outputs").joinpath("yearly_bike.xlsx")
        wb = load_workbook(path_to_inputs)

        # Write data & save
        ws = wb["Data_yearly_stats"]
        print_area = ws["B31:H42"]
        for row_idx, row in enumerate(print_area, 1):
            for column_idx, cell in enumerate(row, 1):
                cell.value = data[row_idx][column_idx]

        wb.save(path_to_outputs)

    def test_busiest_by_various_criteria(self):
        # Import test data pertaining to "mobilité douce"
        installation = models.Installation.objects.get(name="00107695")
        model = models.Model.objects.all().first()
        device = models.Device.objects.all().first()
        sensor_type = models.SensorType.objects.all().first()
        class_ = models.Class.objects.get(name="SPCH-MD 5C")
        tz = pytz.timezone("Europe/Zurich")

        count = models.Count.objects.create(
            start_service_date=tz.localize(datetime(2021, 2, 1)),
            end_service_date=tz.localize(datetime(2021, 12, 10)),
            start_process_date=tz.localize(datetime(2021, 2, 10)),
            end_process_date=tz.localize(datetime(2021, 12, 15)),
            start_put_date=tz.localize(datetime(2021, 1, 1)),
            end_put_date=tz.localize(datetime(2021, 1, 5)),
            id_model=model,
            id_device=device,
            id_sensor_type=sensor_type,
            id_class=class_,
            id_installation=installation,
        )

        path_to_file = Path("/test_data").joinpath(
            "64540060_Latenium_PS2021_ChMixte.txt"
        )
        importer.import_file(str(path_to_file), count)
        print("Imported 1 count files!")

        # Collecting count details
        data = YearlyReportBike.count_details_by_various_criteria(count)

        # Prepare workbook
        path_to_inputs = Path("comptages/report").joinpath("template_yearly_bike.xlsx")
        path_to_outputs = Path("/test_outputs").joinpath("yearly_bike.xlsx")
        wb = load_workbook(path_to_inputs)

        # Write data & save
        ws = wb["Data_yearly_stats"]
        column_names = (
            "VELO",
            "MONO",
            "SHORT",
            "SPECIAL",
            "MULTI",
            "day_or_month_or_weekend",
        )
        row_names = (
            "total_runs_in_year",
            "busiest_date_row",
            "least_busy_date_row",
            "busiest_month_row",
            "least_busy_month_row",
            "total_runs_busiest_hour_weekday",
            "total_runs_busiest_hour_weekend",
        )
        print_area = ws["B2:G8"]
        for row_idx, row_name in enumerate(row_names, 0):
            row = print_area[row_idx]
            YearlyReportBike.write_to_row(
                row_name=row_name,
                row=row,
                data=data,
                key="category_name",
                column_names=column_names,
            )
        wb.save(path_to_outputs)
