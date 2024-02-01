from datetime import datetime
from functools import reduce
import os
from typing import Any, Iterable, Optional, Union
from decimal import Decimal


from django.db.models import Sum, F
from django.db.models.functions import Cast, TruncDate
from django.db.models.fields import DateField
from django.db.models.functions import (
    Cast,
    ExtractHour,
    ExtractIsoWeekDay,
    ExtractMonth,
    TruncDate,
)
from django.db.models import Sum, Avg
from openpyxl import load_workbook

from comptages.core import definitions
from comptages.datamodel.models import CountDetail, Section, Lane, ClassCategory
from comptages.datamodel.models import Count as modelCount


class YearlyReportBike:
    def __init__(self, path_to_output_dir, year, section_id):
        # TODO: pass section or section id?

        self.path_to_output_dir = path_to_output_dir
        self.year = year
        self.section_id = section_id

    def total_runs_by_directions(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            # id_category__code__in=[1, 2],
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Total by day of the week (0->monday, 7->sunday) and by direction
        return (
            qs.annotate(weekday=ExtractIsoWeekDay("timestamp"))
            .values("weekday")
            .annotate(total=Sum("times"))
            .values("weekday", "id_lane__direction", "total")
        )

    def tjms_by_weekday_hour(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # TODO: don't divide by 51 but actually aggregate first by the
        # real days (with sum) and then aggregate by weekday (with average)

        # Total by day of the week (0->monday, 6->sunday) and by hour (0->23)
        result = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(Sum("times"))
            .annotate(weekday=ExtractIsoWeekDay("date"))
            .values("weekday")
            .annotate(tjm=Avg("times"))
            .annotate(hour=ExtractHour("timestamp"))
            .values("weekday", "hour", "tjm")
        )
        return result

    def total_runs_by_hour_and_direction(
        self, directions=(1, 2), weekdays=(0, 1, 2, 3, 4, 5, 6)
    ) -> dict[int, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction__in=directions,
            timestamp__iso_week_day__in=weekdays,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        results = (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(
                runs=Sum("times"),
                direction=F("id_lane__direction"),
                section=F("id_lane__id_section_id"),
            )
            .values("runs", "hour", "direction", "section")
        )

        def partition(acc: dict, val: dict) -> dict:
            hour = val["hour"]
            direction = val["direction"]

            if hour not in acc:
                acc[hour] = {}

            if direction not in acc[hour]:
                acc[hour][direction] = {}

            acc[hour][direction] = val
            return acc

        return reduce(partition, results, {})

    def total_runs_by_hour_one_direction(self, direction: int) -> dict[int, Any]:
        # Get all the count details for hours and the specific direction
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        results = (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(runs=Sum("times"))
            .values("runs", "hour")
            .annotate(day=ExtractIsoWeekDay("timestamp"))
            .order_by("day")
        )

        def reducer(acc: dict, val: dict) -> dict:
            day = val["day"]
            hour = val["hour"]

            if day not in acc:
                acc[day] = {}

            if hour not in acc:
                acc[day][hour] = val["runs"]

            return acc

        return reduce(reducer, results, {})

    def tota_runs_by_hour_weekday_one_direction(
        self, direction: int
    ) -> "ValuesQuerySet[Countdetail, dict[str, Any]]":
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        return (
            qs.annotate(hour=ExtractHour("timestamp"))
            .values("hour")
            .annotate(runs=Sum("times"))
            .values("runs", "hour")
            .annotate(day=ExtractIsoWeekDay("timestamp"))
            .order_by("day")
        )

    def tjms_by_weekday_and_month(
        self,
    ) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # TODO: don't divide by 12 but actually aggregate first by the
        # real days (with sum) and then aggregate by weekday (with average)

        # Total by day of the week (0->monday, 6->sunday) and by month (1->12)
        result = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(Sum("times"))
            .annotate(weekday=ExtractIsoWeekDay("timestamp"))
            .values("weekday")
            .annotate(tjm=Avg("times"), month=ExtractMonth("timestamp"))
            .values("weekday", "month", "tjm")
        )

        return result

    def tjms_by_day(self) -> "ValuesQuerySet[CountDetail, dict[str, Any]]":
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        # Group by date
        result = (
            qs.annotate(date=Cast("timestamp", DateField()))
            .values("date")
            .annotate(tjm=Sum("times"))
            .values("date", "tjm")
        )

        return result

    def tjms_total_runs_by_day_of_week(self) -> dict[str, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        result = (
            qs.annotate(date=TruncDate("timestamp"))
            .values("date")
            .annotate(daily_runs=Sum("times"), week_day=ExtractIsoWeekDay("timestamp"))
            .values("week_day", "daily_runs")
            .order_by("week_day")
        )
        # FIXME
        # Aggregation via `values()` into `annotate()` all the way to the end result would be more performant.
        builder = {}
        for item in result:
            if item["week_day"] not in builder:
                builder[item["week_day"]] = {
                    "days": 1,
                    "runs": item["daily_runs"],
                    "tjm": item["daily_runs"],
                    "week_day": item["week_day"],
                }
            else:
                builder[item["week_day"]]["days"] += 1
                builder[item["week_day"]]["runs"] += item["daily_runs"]
                builder[item["week_day"]]["tjm"] = round(
                    builder[item["week_day"]]["runs"]
                    / builder[item["week_day"]]["days"]
                )
        return builder

    def total_runs_by_class(self) -> dict[str, Any]:
        # Get all the count details for section and the year
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        results = (
            qs.annotate(day=ExtractIsoWeekDay("timestamp"))
            .values("day")
            .annotate(runs=Sum("times"), code=F("id_category__code"))
            .values("day", "runs", "code")
        )

        def reducer(acc: dict, i: dict):
            code = i["code"]
            day = i["day"]
            runs = i["runs"]

            if code not in acc:
                acc[code] = {}

            acc[code][day] = runs
            return acc

        return reduce(reducer, results, {})

    def tjms_by_direction_bike(
        self, categories, direction, weekdays=[0, 1, 2, 3, 4, 5, 6]
    ) -> float:
        qs = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            timestamp__iso_week_day__in=weekdays,
            id_category__code__in=categories,
            id_lane__direction=direction,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        assert qs.exists()

        # TODO: avoid the division?
        return qs.aggregate(res=Sum("times"))["res"] / 365

    def total(self, categories=[1]) -> float:
        qs = CountDetail.objects.filter(
            timestamp__year=self.year,
            id_category__code__in=categories,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )

        return qs.aggregate(res=Sum("times"))["res"]

    def max_day(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(date=Cast("timestamp", DateField()))
            .values("date")
            .annotate(total=Sum("times"))
            .order_by("-total")
        )

        return qs[0]["total"], qs[0]["date"]

    def max_month(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(total=Sum("times"))
            .order_by("-total")
        )

        return qs[0]["total"], qs[0]["month"]

    def min_month(self, categories=[1]) -> tuple[str, Any]:
        qs = (
            CountDetail.objects.filter(
                timestamp__year=self.year,
                id_category__code__in=categories,
                import_status=definitions.IMPORT_STATUS_DEFINITIVE,
            )
            .annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(total=Sum("times"))
            .order_by("total")
        )

        return qs[0]["total"], qs[0]["month"]

    @staticmethod
    def count_details_by_day_month(count: modelCount) -> dict[int, Any]:
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = "SPCH-MD 5C"
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        qs = (
            CountDetail.objects.filter(
                id_count=count.id, id_category__in=categories_ids
            )
            .annotate(
                month=ExtractMonth("timestamp"), day=ExtractIsoWeekDay("timestamp")
            )
            .values("month", "day")
            .annotate(Sum("times"))
        )

        def reducer(acc, item):
            month = item["month"]
            day = item["day"]
            runs = item["times__sum"]
            if month not in acc:
                acc[month] = {}
            if day not in acc[month]:
                acc[month][day] = runs
            return acc

        return reduce(reducer, qs, {})

    @staticmethod
    def count_details_by_various_criteria(
        count: modelCount,
    ) -> dict[str, tuple["ValueQuerySet[CountDetail]", Optional[str]]]:
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = "SPCH-MD 5C"
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        # Base QuerySet
        base_qs = CountDetail.objects.filter(
            id_count=count.id, id_category__in=categories_ids
        )

        # Specialized QuerySets
        total_runs_in_year = (
            base_qs.annotate(category_name=F("id_category__name"))
            .values("category_name")
            .annotate(value=Sum("times"))
        )

        qs = (
            base_qs.annotate(
                category_name=F("id_category__name"), date=TruncDate("timestamp")
            )
            .values("date", "category_name")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )
        busiest_date = qs.first()
        least_busy_date = qs.last()
        assert busiest_date
        assert least_busy_date

        busiest_date_row = (
            base_qs.annotate(
                date=TruncDate("timestamp"), category_name=F("id_category__name")
            )
            .filter(date=busiest_date["date"])
            .values("date", "category_name")
            .annotate(value=Sum("times"))
        )
        least_busy_date_row = (
            base_qs.annotate(
                date=TruncDate("timestamp"), category_name=F("id_category__name")
            )
            .filter(date=least_busy_date["date"])
            .values("date", "category_name")
            .annotate(value=Sum("times"))
        )

        qs = (
            base_qs.annotate(month=ExtractMonth("timestamp"))
            .values("month")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )
        busiest_month = qs.first()
        least_busy_month = qs.last()
        assert busiest_month
        assert least_busy_month

        busiest_month_row = (
            base_qs.annotate(
                month=ExtractMonth("timestamp"), category_name=F("id_category__name")
            )
            .filter(month=busiest_month["month"])
            .values("month", "category_name")
            .annotate(value=Sum("times"))
        )
        least_busy_month_row = (
            base_qs.annotate(
                month=ExtractMonth("timestamp"), category_name=F("id_category__name")
            )
            .filter(month=least_busy_month["month"])
            .values("month", "category_name")
            .annotate(value=Sum("times"))
        )

        qs = (
            base_qs.annotate(
                category_name=F("id_category__name"),
                date=TruncDate("timestamp"),
                hour=ExtractHour("timestamp"),
                week_day=ExtractIsoWeekDay("timestamp"),
            )
            .values("date", "hour", "category_name")
            .annotate(value=Sum("times"))
            .order_by("-value")
        )
        total_runs_busiest_hour_weekday = qs.exclude(week_day__gt=5)
        total_runs_busiest_hour_weekend = qs.exclude(week_day__lt=6)

        busiest_weekday = total_runs_busiest_hour_weekday.first()
        busiest_weekend = total_runs_busiest_hour_weekend[:2]
        assert busiest_weekday
        assert busiest_weekend

        return {
            "busiest_date_row": (busiest_date_row, busiest_date["date"]),
            "least_busy_date_row": (least_busy_date_row, str(least_busy_date["date"])),
            "busiest_month_row": (busiest_month_row, str(busiest_month["month"])),
            "least_busy_month_row": (
                least_busy_month_row,
                str(least_busy_month["month"]),
            ),
            "total_runs_busiest_hour_weekday": (
                total_runs_busiest_hour_weekday,
                str(busiest_weekday["date"]),
            ),
            "total_runs_busiest_hour_weekend": (
                total_runs_busiest_hour_weekend,
                ", ".join(str(item["date"]) for item in busiest_weekend),
            ),
            "total_runs_in_year": (total_runs_in_year, None),
        }

    @staticmethod
    def count_details_by_season(count: modelCount) -> dict[int, Any]:
        """Break down count details by season x section x class"""
        # Assuming seasons to run from 20 <month> to 21 <month n + 1>
        seasons = {
            "printemps": [3, 4, 5],
            "été": [6, 7, 8],
            "automne": [9, 10, 11],
            "hiver": [12, 1, 2],
        }
        # Preparing to filter out categories that don't reference the class picked out by `class_name`
        class_name = "SPCH-MD 5C"
        # Excluding irrelevant
        categories_name_to_exclude = ("TRASH", "ELSE")
        categories_ids = (
            ClassCategory.objects.filter(id_class__name=class_name)
            .exclude(id_category__name__in=categories_name_to_exclude)
            .values_list("id_category", flat=True)
        )
        # Getting data
        count_details = (
            CountDetail.objects.filter(
                id_count=count.id, id_category__in=categories_ids
            )
            .annotate(
                section=F("id_lane__id_section"), category_name=F("id_category__name")
            )
            .values("id", "section", "category_name", "times", "timestamp")
        )

        # Preparing to collect data
        def reducer(acc: dict, detail) -> dict:
            timestamp: datetime = detail["timestamp"]

            for season, _range in seasons.items():
                if timestamp.month in _range and (
                    timestamp.month != _range[0] or timestamp.day >= 21
                ):
                    section_id = detail["section"]
                    category_name = detail["category_name"]
                    times = detail["times"]

                    if season not in acc:
                        acc[season] = {}

                    if category_name not in acc[season]:
                        acc[season][category_name] = {}

                    if section_id not in acc[season][category_name]:
                        acc[season][category_name][section_id] = 0

                    acc[season][category_name][section_id] += times
                    break

            return acc

        # Collecting
        return reduce(reducer, count_details, {})

    @staticmethod
    def write_to_row(
        *,
        row_name: str,
        row: Iterable,
        data: dict,
        key: str,
        column_names: Iterable[str],
    ):
        items, day_or_month_or_weekend = data[row_name]
        for column_name, cell in zip(column_names, row):
            if column_name == "day_or_month_or_weekend":
                cell.value = day_or_month_or_weekend or "-"
            elif item := next(
                filter(
                    lambda item: (item[key] == column_name),
                    items,
                ),
                None,
            ):
                cell.value = item["value"]
            else:
                cell.value = "-"

    def run(self):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        template = os.path.join(current_dir, "template_yearly_bike.xlsx")
        workbook = load_workbook(filename=template)

        """ Data_count """

        ws = workbook["Data_count"]

        section = Section.objects.get(id=self.section_id)

        def render_section_dist(value: Union[str, Decimal, None]) -> str:
            if value is None or value == "NA":
                return "NA"
            if isinstance(value, str):
                return str(round(int(value)))
            if isinstance(value, Decimal):
                return str(round(value))
            raise ValueError(value)

        section_start_dist = render_section_dist(section.start_dist)
        section_end_dist = render_section_dist(section.end_dist)

        ws[
            "B3"
        ] = f"""
            Poste de comptage : {section.id}  
            Axe : {section.owner}:{section.road}{section.way}
            PR {section.start_pr} + {section_start_dist} m à PR {section.end_pr} + {section_end_dist} m
        """

        ws["B4"] = "Periode de comptage du 01/01/{0} au 31/12/{0}".format(self.year)
        ws["B5"] = "Comptage {}".format(self.year)

        # Get one count for the section and the year to get the base data
        count_detail = CountDetail.objects.filter(
            id_lane__id_section__id=self.section_id,
            timestamp__year=self.year,
            import_status=definitions.IMPORT_STATUS_DEFINITIVE,
        )
        if not count_detail.exists():
            print(
                "Aucun conmptage pour cette année ({}) et cette section ({})".format(
                    self.year, self.section_id
                )
            )
            return

        count = count_detail[0].id_count

        ws["B6"] = "Type de capteur : {}".format(count.id_sensor_type.name)
        ws["B7"] = "Modèle : {}".format(count.id_model.name)
        ws["B8"] = "Classification : {}".format(count.id_class.name)
        ws["B9"] = "Comptage véhicule par véhicule"
        ws["B12"] = "Remarque : {}".format(count.remarks)

        lanes = Lane.objects.filter(id_installation=count.id_installation)

        ws["B13"] = lanes[0].direction_desc
        if len(lanes) > 1:
            ws["B14"] = lanes[1].direction_desc

        ws["B11"] = lanes[0].id_section.place_name

        """ AN_TE"""

        ws = workbook["AN_TE"]

        row_offset = 14
        column_offset = 1
        data = self.tjms_by_weekday_hour()
        for i in data:
            ws.cell(
                row=i["hour"] + row_offset,
                column=i["weekday"] + column_offset,
                value=i["tjm"],
            )

        row_offset = 47
        column_offset = 1

        data = self.tjms_by_weekday_and_month()
        row = row_offset
        for i in data:
            ws.cell(row=row, column=column_offset, value=i["date"])
            ws.cell(row=row, column=column_offset + 1, value=i["tjm"])
            row += 1

        """ CV_LV """

        ws = workbook["CV_LV"]

        ws["F12"] = self.tjms_by_direction_bike([1], 1)
        ws["G12"] = self.tjms_by_direction_bike([1], 2)
        ws["H12"] = self.tjms_by_direction_bike([2, 3, 4, 5], 1)
        ws["I12"] = self.tjms_by_direction_bike([2, 3, 4, 5], 2)

        ws["J35"] = self.total()
        ws["J39"] = self.max_day()[0]
        ws["K39"] = self.max_day()[1]

        ws["J40"] = self.max_month()[0]
        ws["K40"] = self.max_month()[1]

        ws["J41"] = self.min_month()[0]
        ws["k41"] = self.min_month()[1]

        """ Data_year """

        ws = workbook["Data_year"]
        row_offset = 4
        column_offset = 1

        data = self.tjms_by_day()
        row = row_offset
        for i in data:
            ws.cell(row=row, column=column_offset, value=i["date"])
            ws.cell(row=row, column=column_offset + 1, value=i["tjm"])
            row += 1

        """ Data_week """

        ws = workbook["Data_week"]
        row_offset = 4
        column_offset = 2

        data = self.tjms_total_runs_by_day_of_week().values()
        row = row_offset
        for i in data:
            ws.cell(row=row, column=column_offset, value=i["runs"])
            ws.cell(row=row, column=column_offset + 3, value=i["tjm"])
            row += 1

        """ Data_hour """

        ws = workbook["Data_hour"]

        # Data hour > Whole weeks
        print_area = ws["C5:D28"]
        data = self.total_runs_by_hour_and_direction(directions=(1, 2))
        for hour, row in enumerate(print_area, 1):
            if hour == 24:
                hour = 0
            for direction, cell in enumerate(row, 1):
                if hour not in data or direction not in data[hour]:
                    cell.value = 0
                else:
                    cell.value = data[hour][direction]["runs"]

        # Data hour > Weekends only
        print_area = ws["C37:D60"]
        data = self.total_runs_by_hour_and_direction(directions=(1, 2), weekdays=(5, 6))
        for hour, row in enumerate(print_area, 1):
            if hour == 24:
                hour = 0
            for direction, cell in enumerate(row, 1):
                if hour not in data or direction not in data[hour]:
                    cell.value = 0
                else:
                    cell.value = data[hour][direction]["runs"]

        # Data hour > Only dir 1
        print_area = ws["B69:H92"]
        data = self.total_runs_by_hour_one_direction(1)
        for hour, row in enumerate(print_area, 1):
            if hour == 24:
                hour = 0
            for day, cell in enumerate(row, 1):
                if day not in data or hour not in data[day]:
                    cell.value = 0
                else:
                    cell.value = data[day][hour]

        # Data hour > Only dir 2
        print_area = ws["B101:H124"]
        data = self.total_runs_by_hour_one_direction(2)
        for hour, row in enumerate(print_area, 1):
            if hour == 24:
                hour = 0
            for day, cell in enumerate(row, 1):
                if day not in data or hour not in data[day]:
                    cell.value = 0
                else:
                    cell.value = data[day][hour]

        """ Data_yearly_stats """

        ws = workbook["Data_yearly_stats"]
        print_area = ws["B2:G8"]
        data = YearlyReportBike.count_details_by_various_criteria(count)
        column_names = (
            "VELO",
            "MONO",
            "SHORT",
            "SPECIAL",
            "MULTI",
            "day_or_month_or_weekend",
        )
        row_names = (
            "total_runs_in_year",
            "busiest_date_row",
            "least_busy_date_row",
            "busiest_month_row",
            "least_busy_month_row",
            "total_runs_busiest_hour_weekday",
            "total_runs_busiest_hour_weekend",
        )
        for row_idx, row_name in enumerate(row_names, 0):
            row = print_area[row_idx]
            YearlyReportBike.write_to_row(
                row_name=row_name,
                row=row,
                data=data,
                key="category_name",
                column_names=column_names,
            )

        """ Data_class """

        ws = workbook["Data_class"]
        print_area = ws["B4:H18"]
        for code, row in enumerate(print_area, 0):
            for day, cell in enumerate(row, 1):
                if code not in data or day not in data[code]:
                    cell.value = 0
                else:
                    cell.value = data[code][day]

        ws = workbook["AN_GR"]
        ws.print_area = "A1:Z62"

        ws = workbook["CAT"]
        ws.print_area = "A1:Z62"

        # Save the file
        output = os.path.join(
            self.path_to_output_dir, "{}_{}_r.xlsx".format(self.section_id, self.year)
        )
        workbook.save(filename=output)
        print(f"Saved report to {output}")
